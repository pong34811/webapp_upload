"""
Excel import helpers for .xlsx files.

Expected columns in each sheet (first row = header):
  - filename   (required)  – video file name (matched from uploaded files)
  - title      (optional)  – YouTube / Facebook video title
  - description (optional) – description / caption
  - tags       (optional)  – comma-separated tags
  - privacy    (optional)  – public | private | unlisted
  - scheduled_time (optional) – ISO-8601 datetime string
"""

import io
import os
from typing import Any

from openpyxl import load_workbook

# Column alias map: canonical name → set of accepted header variants
COLUMN_ALIASES: dict[str, set[str]] = {
    "filename":        {"filename", "video_path", "file", "file_path", "video"},
    "title":           {"title", "video_title", "name"},
    "description":     {"description", "desc"},
    "tags":            {"tags", "tag", "keyword", "keywords"},
    "privacy":         {"privacy", "privacystatus", "privacy_status", "visibility"},
    "scheduled_time":  {"scheduled_time", "publishat", "publish_at", "scheduled", "schedule"},
    "thumbnail_path":  {"thumbnail_path", "thumbnail", "thumb"},
    "categoryid":      {"categoryid", "category_id"},
    "playlistid":      {"playlistid", "playlist_id", "playlist"},
    "state":           {"state", "status"},
}

# Build reverse lookup: variant → canonical
_VARIANT_TO_CANONICAL: dict[str, str] = {}
for _canon, _variants in COLUMN_ALIASES.items():
    for _v in _variants:
        _VARIANT_TO_CANONICAL[_v] = _canon

# The importer requires at least a video file reference
REQUIRED_COLUMNS = {"filename"}  # canonical names
OPTIONAL_COLUMNS = {"title", "description", "tags", "privacy",
                    "scheduled_time", "thumbnail_path", "categoryid",
                    "playlistid", "state"}
ALL_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS


def _normalise(col: str) -> str:
    """Lower-case, strip, and map to canonical column name."""
    raw = col.strip().lower().replace(" ", "_")
    return _VARIANT_TO_CANONICAL.get(raw, raw)


def _parse_rows(ws) -> tuple[list[str], list[dict[str, Any]]]:
    """Read an openpyxl worksheet into headers + row dicts."""
    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_header = next(rows_iter)
    except StopIteration:
        return [], []

    headers = [_normalise(h) if h else "" for h in raw_header]
    # drop trailing empty headers
    while headers and not headers[-1]:
        headers.pop()

    data: list[dict[str, Any]] = []
    for row in rows_iter:
        if not any(cell is not None for cell in row[: len(headers)]):
            continue  # skip blank rows
        record: dict[str, Any] = {}
        for idx, h in enumerate(headers):
            if not h:
                continue
            value = row[idx] if idx < len(row) else None
            # Normalise datetime objects to ISO string
            if hasattr(value, "isoformat"):
                value = value.isoformat()
            elif value is not None:
                value = str(value).strip()
            record[h] = value
        data.append(record)
    return headers, data


def parse_excel(file_bytes: bytes) -> dict[str, Any]:
    """
    Parse an uploaded .xlsx file and return sheet names + their headers.

    Returns:
        {
            "sheets": [
                {"name": "Sheet1", "columns": ["filename", "title", ...]},
                ...
            ]
        }
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets_info = []
    for title in wb.sheetnames:
        ws = wb[title]
        headers, _ = _parse_rows(ws)
        sheets_info.append({"name": title, "columns": headers})
    wb.close()
    return {"sheets": sheets_info}


def preview_sheet(file_bytes: bytes, sheet_name: str) -> dict[str, Any]:
    """
    Parse a specific sheet and return preview rows (first 50) plus validation info.

    Returns:
        {
            "sheet": "Sheet1",
            "columns": [...],
            "total_rows": 42,
            "rows": [
                {
                    "row": 2,
                    "filename": "video1.mp4",
                    "title": "...",
                    "errors": []
                },
                ...
            ]
        }
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]
    headers, rows = _parse_rows(ws)
    wb.close()

    missing_cols = REQUIRED_COLUMNS - set(headers)
    if missing_cols:
        raise ValueError(
            f"Missing required columns: {', '.join(sorted(missing_cols))}"
        )

    validated: list[dict[str, Any]] = []
    skipped_non_upload = 0
    for i, row in enumerate(rows, start=2):  # Excel row 1 = header
        # Only import rows where state == WAIT_UPLOAD
        state_val = (row.get("state") or "").strip().upper()
        if state_val and state_val != "WAIT_UPLOAD":
            skipped_non_upload += 1
            continue

        errors: list[str] = []
        fn = row.get("filename")
        if not fn:
            errors.append("filename is required")
        priv = row.get("privacy")
        if priv and priv.lower() not in ("public", "private", "unlisted"):
            errors.append(f"invalid privacy '{priv}' (use public/private/unlisted)")
        raw_fn = row.get("filename") or ""
        # Extract basename from full path (e.g. "G:\\folder\\video.mp4" → "video.mp4")
        basename = os.path.basename(raw_fn.replace("\\", "/")) if raw_fn else ""
        # Check if the video file actually exists on disk
        file_found = bool(raw_fn) and os.path.isfile(raw_fn)
        validated.append({
            "row": i,
            "filename": basename or raw_fn,
            "filename_full": raw_fn,
            "file_found": file_found,
            "title": row.get("title") or "",
            "description": row.get("description") or "",
            "tags": row.get("tags") or "",
            "privacy": row.get("privacy") or "private",
            "scheduled_time": row.get("scheduled_time") or "",
            "state": state_val,
            "errors": errors,
        })

    return {
        "sheet": sheet_name,
        "columns": headers,
        "total_rows": len(validated),
        "skipped_non_upload": skipped_non_upload,
        "rows": validated,
    }
